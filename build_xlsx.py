# -*- coding: utf-8 -*-
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

wb = Workbook()

BLUE = Font(name="Microsoft JhengHei", color="0000FF", size=11)
BLACK = Font(name="Microsoft JhengHei", color="000000", size=11)
BOLD = Font(name="Microsoft JhengHei", bold=True, size=11)
TITLE = Font(name="Microsoft JhengHei", bold=True, size=16, color="0E7490")
HDR_FILL = PatternFill("solid", start_color="0E7490")
HDR_FONT = Font(name="Microsoft JhengHei", bold=True, color="FFFFFF", size=11)
YELLOW = PatternFill("solid", start_color="FFFF00")
LIGHT = PatternFill("solid", start_color="E0F2FE")
thin = Side(style="thin", color="B0BEC5")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
CENTER = Alignment(horizontal="center", vertical="center")

# ── Sheet 2: 單價參考 ──
ref = wb.create_sheet("單價參考")
ref_data = [
    ["目的地", "機票來回(TWD/人)", "住宿(TWD/人/晚)", "餐飲(TWD/人/日)", "當地交通(TWD/人/日)", "門票活動(TWD/人/日)", "保險(TWD/人/日)", "購物預算(TWD/人)", "匯率(1外幣=?TWD)", "幣別"],
    ["台灣", 0, 1600, 600, 600, 300, 60, 1000, 1, "TWD"],
    ["日本", 12000, 2200, 1200, 700, 500, 150, 3000, 0.21, "JPY"],
    ["韓國", 9000, 1800, 1000, 500, 400, 150, 3000, 0.023, "KRW"],
    ["泰國", 9500, 1300, 700, 450, 600, 180, 3000, 0.95, "THB"],
]
for r, row in enumerate(ref_data, 1):
    for c, v in enumerate(row, 1):
        cell = ref.cell(row=r, column=c, value=v)
        cell.border = BORDER
        if r == 1:
            cell.fill = HDR_FILL; cell.font = HDR_FONT; cell.alignment = CENTER
        elif c > 1:
            cell.font = BLUE
        else:
            cell.font = BOLD
ref.cell(row=7, column=1, value="※ 單價為估算值（藍字可自行修改）。來源：2026 年市場行情概估，請依實際報價調整。").font = Font(name="Microsoft JhengHei", italic=True, size=9, color="808080")
for col, w in zip("ABCDEFGHIJ", [10, 18, 18, 18, 20, 20, 16, 18, 18, 8]):
    ref.column_dimensions[col].width = w

# ── Sheet 1: 預算試算 ──
ws = wb.active
ws.title = "預算試算"
ws["A1"] = "🧳 旅遊預算試算表"
ws["A1"].font = TITLE
ws["A2"] = "藍字＝可修改的輸入值；黑字＝公式自動計算。改目的地/人數/天數即時重算。"
ws["A2"].font = Font(name="Microsoft JhengHei", size=9, color="808080")

inputs = [("目的地", "日本"), ("出發日期", "2026-07-15"), ("天數", 5), ("大人人數", 2), ("小孩人數(半價計)", 0), ("風格係數(0.8省/1標準/1.5豪華)", 1)]
for i, (label, val) in enumerate(inputs, 4):
    ws.cell(row=i, column=1, value=label).font = BOLD
    c = ws.cell(row=i, column=2, value=val)
    c.font = BLUE; c.fill = YELLOW; c.border = BORDER; c.alignment = CENTER

dv_dest = DataValidation(type="list", formula1='"台灣,日本,韓國,泰國"', allow_blank=False)
dv_style = DataValidation(type="list", formula1='"0.8,1,1.5"', allow_blank=False)
ws.add_data_validation(dv_dest); ws.add_data_validation(dv_style)
dv_dest.add("B4"); dv_style.add("B9")

ws["D4"] = "有效人數(小孩半價)"; ws["D4"].font = BOLD
ws["E4"] = "=B7+B8*0.5"; ws["E4"].font = BLACK; ws["E4"].border = BORDER; ws["E4"].alignment = CENTER
ws["D5"] = "住宿晚數"; ws["D5"].font = BOLD
ws["E5"] = "=MAX(B6-1,0)"; ws["E5"].font = BLACK; ws["E5"].border = BORDER; ws["E5"].alignment = CENTER
ws["D6"] = "匯率"; ws["D6"].font = BOLD
ws["E6"] = '=VLOOKUP($B$4,單價參考!$A$2:$J$5,9,FALSE)'; ws["E6"].font = BLACK; ws["E6"].border = BORDER; ws["E6"].alignment = CENTER
ws["D7"] = "幣別"; ws["D7"].font = BOLD
ws["E7"] = '=VLOOKUP($B$4,單價參考!$A$2:$J$5,10,FALSE)'; ws["E7"].font = BLACK; ws["E7"].border = BORDER; ws["E7"].alignment = CENTER

# 預算表
hdr_row = 12
headers = ["項目", "計算方式", "單價(TWD)", "數量", "小計(TWD)", "外幣參考"]
for c, h in enumerate(headers, 1):
    cell = ws.cell(row=hdr_row, column=c, value=h)
    cell.fill = HDR_FILL; cell.font = HDR_FONT; cell.alignment = CENTER; cell.border = BORDER

# (label, mode, vlookup col, style-applied?, qty formula)
rows = [
    ("機票（來回）", "每人一次", 2, False, "=$E$4"),
    ("住宿", "每人每晚", 3, True, "=$E$4*$E$5"),
    ("餐飲", "每人每日", 4, True, "=$E$4*$B$6"),
    ("當地交通", "每人每日", 5, True, "=$E$4*$B$6"),
    ("景點門票／活動", "每人每日", 6, True, "=$E$4*$B$6"),
    ("旅遊保險", "每人每日", 7, False, "=$E$4*$B$6"),
    ("購物／伴手禮", "每人一次", 8, False, "=$E$4"),
]
r0 = hdr_row + 1
for i, (label, mode, col, styled, qty) in enumerate(rows):
    r = r0 + i
    ws.cell(row=r, column=1, value=label).font = BOLD
    ws.cell(row=r, column=2, value=mode).font = BLACK
    price = f'=VLOOKUP($B$4,單價參考!$A$2:$J$5,{col},FALSE)' + ("*$B$9" if styled else "")
    ws.cell(row=r, column=3, value=price).font = BLACK
    ws.cell(row=r, column=4, value=qty).font = BLACK
    ws.cell(row=r, column=5, value=f"=C{r}*D{r}").font = BLACK
    ws.cell(row=r, column=6, value=f'=IF($E$6>0,E{r}/$E$6,0)').font = BLACK
    for c in range(1, 7):
        ws.cell(row=r, column=c).border = BORDER

last = r0 + len(rows) - 1
sub_r, res_r, tot_r, pp_r, fx_r = last + 1, last + 2, last + 3, last + 4, last + 5
ws.cell(row=sub_r, column=1, value="小計").font = BOLD
ws.cell(row=sub_r, column=5, value=f"=SUM(E{r0}:E{last})").font = BOLD
ws.cell(row=res_r, column=1, value="預備金 10%").font = BOLD
ws.cell(row=res_r, column=5, value=f"=E{sub_r}*0.1").font = BLACK
ws.cell(row=tot_r, column=1, value="總預算").font = Font(name="Microsoft JhengHei", bold=True, size=13, color="0E7490")
ws.cell(row=tot_r, column=5, value=f"=E{sub_r}+E{res_r}").font = Font(name="Microsoft JhengHei", bold=True, size=13, color="0E7490")
ws.cell(row=pp_r, column=1, value="每人平均(含小孩)").font = BOLD
ws.cell(row=pp_r, column=5, value=f"=IF(B7+B8>0,E{tot_r}/(B7+B8),0)").font = BLACK
ws.cell(row=fx_r, column=1, value="總預算外幣參考").font = BOLD
ws.cell(row=fx_r, column=5, value=f"=IF($E$6>0,E{tot_r}/$E$6,0)").font = BLACK
for rr in (sub_r, res_r, tot_r, pp_r, fx_r):
    for c in range(1, 7):
        ws.cell(row=rr, column=c).border = BORDER
    ws.cell(row=rr, column=5).fill = LIGHT

money = '#,##0;(#,##0);"-"'
for r in range(r0, fx_r + 1):
    for c in (3, 5, 6):
        ws.cell(row=r, column=c).number_format = money
for col, w in zip("ABCDEF", [22, 14, 14, 14, 16, 14]):
    ws.column_dimensions[col].width = w

# ── Sheet 3: 行程表 ──
it = wb.create_sheet("行程表")
it["A1"] = "🗓️ 逐日行程表"; it["A1"].font = TITLE
hdrs = ["第幾天", "日期", "時間", "行程內容", "交通方式", "預估費用(TWD)", "備註"]
for c, h in enumerate(hdrs, 1):
    cell = it.cell(row=3, column=c, value=h)
    cell.fill = HDR_FILL; cell.font = HDR_FONT; cell.alignment = CENTER; cell.border = BORDER
sample = [
    [1, "=預算試算!B5", "10:00", "抵達機場、領 WiFi/eSIM", "機場快線", 300, "範例列，請自行修改"],
    [1, "=預算試算!B5", "15:00", "飯店 Check-in、周邊散步", "地鐵", 50, ""],
    [2, "", "09:00", "（自行填入景點）", "", "", ""],
]
for r, row in enumerate(sample, 4):
    for c, v in enumerate(row, 1):
        cell = it.cell(row=r, column=c, value=v)
        cell.border = BORDER
        cell.font = BLACK if (c == 2 and r <= 5) else BLUE
for r in range(7, 24):
    for c in range(1, 8):
        it.cell(row=r, column=c).border = BORDER
it.cell(row=25, column=5, value="費用合計").font = BOLD
it.cell(row=25, column=6, value="=SUM(F4:F23)").font = BLACK
it.cell(row=25, column=6).number_format = money
for col, w in zip("ABCDEFG", [8, 12, 8, 36, 14, 16, 24]):
    it.column_dimensions[col].width = w

# ── Sheet 4: 攜帶清單 ──
pk = wb.create_sheet("攜帶清單")
pk["A1"] = "🎒 攜帶物品檢查清單"; pk["A1"].font = TITLE
pk["A2"] = "「已備」欄填 1 表示完成；進度自動計算。"
pk["A2"].font = Font(name="Microsoft JhengHei", size=9, color="808080")
for c, h in enumerate(["分類", "物品", "適用", "已備(填1)"], 1):
    cell = pk.cell(row=4, column=c, value=h)
    cell.fill = HDR_FILL; cell.font = HDR_FONT; cell.alignment = CENTER; cell.border = BORDER
items = [
    ("必帶", "護照（效期6個月以上）", "出國"), ("必帶", "身分證／健保卡", "全部"),
    ("必帶", "手機＋充電線＋行動電源", "全部"), ("必帶", "現金（台幣＋外幣）", "全部"),
    ("必帶", "信用卡（開通海外刷卡）", "出國"), ("必帶", "網路（eSIM／WiFi機）", "出國"),
    ("必帶", "常用藥品＋腸胃藥", "全部"), ("必帶", "訂房／機票證明（紙本＋截圖）", "全部"),
    ("必帶", "旅遊平安險保單", "出國"), ("必帶", "轉接頭（220V圓孔）", "韓國/泰國"),
    ("必帶", "Visit Japan Web QR Code", "日本"), ("必帶", "TDAC 數位入境卡", "泰國"),
    ("建議", "折疊傘／雨衣", "全部"), ("建議", "防曬乳＋帽子", "全部"),
    ("建議", "好走的鞋", "全部"), ("建議", "保溫瓶", "全部"),
    ("建議", "空行李袋（戰利品）", "出國"), ("建議", "防蚊液", "台灣/泰國"),
    ("建議", "發熱衣／暖暖包（冬季）", "日本/韓國"), ("建議", "手機防水袋", "泰國"),
    ("建議", "護照影本（與正本分開放）", "出國"), ("建議", "緊急聯絡人資訊卡", "全部"),
]
for r, (cat, name, scope) in enumerate(items, 5):
    pk.cell(row=r, column=1, value=cat).font = Font(name="Microsoft JhengHei", bold=True, color="B91C1C" if cat == "必帶" else "B45309")
    pk.cell(row=r, column=2, value=name).font = BLACK
    pk.cell(row=r, column=3, value=scope).font = BLACK
    pk.cell(row=r, column=4).font = BLUE
    for c in range(1, 5):
        pk.cell(row=r, column=c).border = BORDER
last_i = 4 + len(items)
pk.cell(row=last_i + 2, column=2, value="完成進度").font = BOLD
prog = pk.cell(row=last_i + 2, column=3, value=f'=COUNTIF(D5:D{last_i},1)&" / "&COUNTA(B5:B{last_i})&" 項"')
prog.font = BLACK
for col, w in zip("ABCD", [8, 36, 14, 10]):
    pk.column_dimensions[col].width = w

wb.save(r"C:\Users\zhanh\OneDrive\桌面\AI練習區\旅遊規劃系統\旅遊預算試算表.xlsx")
print("OK")
